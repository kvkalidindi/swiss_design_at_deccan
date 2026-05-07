"""Build deccan.xltx using openpyxl.

Excel template with brand "furniture":
  1. Cover sheet  — first sheet, logo + title + metadata block
  2. Print header — small logo (left) + workbook title (right) on every printed page
  3. Print footer — confidentiality line (left) + page number (right) on every printed page
  4. End sheet    — last sheet, centered logo + brand line

See skill/references/document-furniture.md for the canonical spec.
"""
from __future__ import annotations

import json
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from scripts.lib.office_theme import build_theme_xml

ROOT = Path(__file__).resolve().parents[2]
PALETTE = ROOT / "outputs" / "palette.json"
LOGO = ROOT / "data" / "logo.png"
OUT = ROOT / "office" / "templates" / "deccan.xltx"


def _fit_to_one_page(ws) -> None:
    """Force a sheet to print on a single page so cover/end/data don't spill
    onto blank trailing pages with stray accent rules or truncated text."""
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _set_print_header_footer(ws, title_text: str, classification: str = "Confidential") -> None:
    """Configure Excel print header/footer with logo + title and classification + page #."""
    # Header: left = logo (&G), right = title in 9pt IBM Plex Sans
    ws.oddHeader.left.text = "&G"
    ws.oddHeader.left.size = 9
    ws.oddHeader.left.font = "IBM Plex Sans"
    ws.oddHeader.right.text = title_text
    ws.oddHeader.right.size = 9
    ws.oddHeader.right.font = "IBM Plex Sans"
    ws.oddHeader.right.color = "595959"

    # Footer: left = "Deccan Fine Chemicals · classification", right = "Page X of Y"
    ws.oddFooter.left.text = f"Deccan Fine Chemicals · {classification}"
    ws.oddFooter.left.size = 9
    ws.oddFooter.left.font = "IBM Plex Sans"
    ws.oddFooter.left.color = "595959"
    ws.oddFooter.right.text = "&P"
    ws.oddFooter.right.size = 9
    ws.oddFooter.right.font = "IBM Plex Sans"
    ws.oddFooter.right.color = "595959"

    if LOGO.exists():
        # Embed logo into the print header's left section (&G placeholder)
        try:
            ws.HeaderFooter.differentFirst = False
            ws.HeaderFooter.differentOddEven = False
        except AttributeError:
            pass


def _build_cover_sheet(wb, blue_500: str) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    # Anchor the logo in the top area
    if LOGO.exists():
        try:
            img = XLImage(str(LOGO))
            # Constrain logo width ~ 240px (matches ~2.5" at 96 DPI)
            img.width = 240
            img.height = int(img.height * (240 / max(img.width, 1)))
            ws.add_image(img, "B2")
        except Exception:
            pass

    # Title (row 8)
    ws["B8"] = "Document Title"
    ws["B8"].font = Font(name="IBM Plex Sans", size=36, color=blue_500, bold=False)
    ws["B8"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[8].height = 50
    ws.merge_cells("B8:H8")

    # Subtitle (row 9)
    ws["B9"] = "Subtitle / one-line summary"
    ws["B9"].font = Font(name="IBM Plex Sans", size=14, color="595959")
    ws["B9"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[9].height = 24
    ws.merge_cells("B9:H9")

    # Accent rule via a thin filled row (row 10)
    rule_fill = PatternFill("solid", fgColor=blue_500)
    for col in range(2, 9):
        ws.cell(row=10, column=col).fill = rule_fill
    ws.row_dimensions[10].height = 3

    # Metadata block (rows 12..16)
    labels = ["DOCUMENT TYPE", "PREPARED BY", "DATE", "VERSION", "CLASSIFICATION"]
    values = ["Workbook", "Author Name", "YYYY-MM-DD", "v1.0", "Confidential"]
    label_font = Font(name="IBM Plex Sans", size=8, color="8A8786")
    value_font = Font(name="IBM Plex Sans", size=10, color="1C1917")
    for i, (label, value) in enumerate(zip(labels, values)):
        r = 12 + i
        ws.cell(row=r, column=2, value=label).font = label_font
        ws.cell(row=r, column=3, value=value).font = value_font
        ws.row_dimensions[r].height = 16

    # Sensible column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    for letter in "CDEFGH":
        ws.column_dimensions[letter].width = 18

    # Set the cover's print area to the live composition only — empty rows
    # below the metadata block must not paginate onto a blank "page 2".
    ws.print_area = "A1:H20"
    _fit_to_one_page(ws)


def _build_data_sheet(wb, blue_500: str, blue_700: str, stone_50: str) -> None:
    ws = wb.create_sheet("Sheet1")

    ws["A1"] = "Document Title"
    ws["A1"].font = Font(name="IBM Plex Sans", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=blue_500)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:F1")

    ws["A3"] = "Section Header"
    ws["A3"].font = Font(name="IBM Plex Sans", size=13, bold=True, color=blue_700)

    headers = ["Column A", "Column B", "Column C", "Column D"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = Font(name="IBM Plex Sans", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue_500)
        cell.alignment = Alignment(horizontal="left")

    for row_idx in range(6, 16):
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.font = Font(name="IBM Plex Sans", size=10)
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=stone_50)

    for col_letter, width in zip("ABCDEF", [22, 18, 18, 18, 18, 18]):
        ws.column_dimensions[col_letter].width = width

    _set_print_header_footer(ws, "Document Title")
    ws.print_area = "A1:F30"
    _fit_to_one_page(ws)


def _build_end_sheet(wb) -> None:
    ws = wb.create_sheet("End")
    ws.sheet_view.showGridLines = False

    if LOGO.exists():
        try:
            img = XLImage(str(LOGO))
            img.width = 200
            img.height = int(img.height * (200 / max(img.width, 1)))
            ws.add_image(img, "D6")
        except Exception:
            pass

    ws["A14"] = "Deccan Fine Chemicals"
    ws["A14"].font = Font(name="IBM Plex Sans", size=16, color="1C1917")
    ws["A14"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A14:H14")
    ws.row_dimensions[14].height = 24

    ws["A15"] = "deccanchemicals.com · Hyderabad, India"
    ws["A15"].font = Font(name="IBM Plex Sans", size=10, color="8A8786")
    ws["A15"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A15:H15")
    ws.row_dimensions[15].height = 16

    for letter in "ABCDEFGH":
        ws.column_dimensions[letter].width = 14

    ws.print_area = "A1:H20"
    _fit_to_one_page(ws)


def emit_xltx() -> Path:
    palette = json.loads(PALETTE.read_text(encoding="utf-8"))
    blue_500 = palette["blue"]["500"]["hex"].lstrip("#")
    blue_700 = palette["blue"]["700"]["hex"].lstrip("#")
    stone_50 = "FAFAF9"

    wb = openpyxl.Workbook()
    # Drop the default sheet; we'll create our own in order.
    default_ws = wb.active
    wb.remove(default_ws)

    _build_cover_sheet(wb, blue_500)
    _build_data_sheet(wb, blue_500, blue_700, stone_50)
    _build_end_sheet(wb)

    # Make Cover the active sheet on open
    wb.active = wb.sheetnames.index("Cover")

    tmp = OUT.with_suffix(".xlsx")
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(tmp)

    _replace_theme_in_xlsx(tmp, build_theme_xml(palette))
    _convert_to_xltx(tmp, OUT)
    tmp.unlink()
    return OUT


def _replace_theme_in_xlsx(path: Path, theme_xml: str) -> None:
    buf = BytesIO()
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            theme_replaced = False
            for item in zin.namelist():
                data = zin.read(item)
                if item == "xl/theme/theme1.xml":
                    data = theme_xml.encode("utf-8")
                    theme_replaced = True
                zout.writestr(item, data)
            if not theme_replaced:
                zout.writestr("xl/theme/theme1.xml", theme_xml)
    path.write_bytes(buf.getvalue())


def _convert_to_xltx(src: Path, dst: Path) -> None:
    buf = BytesIO()
    with zipfile.ZipFile(src, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    text = text.replace(
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
                    )
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    dst.write_bytes(buf.getvalue())
